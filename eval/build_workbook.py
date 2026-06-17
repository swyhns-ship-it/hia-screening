# -*- coding: utf-8 -*-
"""③ 合并 Opus 预判 + 生成人工复核工作簿(xlsx)。
- 1-160(重复)未在 verdicts 中的,补 drop/agree=C(逐条核过,全属真冗余)。
- 输出 eval/adjudication_workbook.xlsx:flag/分歧行置顶高亮,其余按模式排;留空列给人工判。
"""
import io
import json
import os
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JL = os.path.join(ROOT, "eval", "adjudication_272.jsonl")
VJ = os.path.join(ROOT, "eval", "adjudication_verdicts.json")
XLSX = os.path.join(ROOT, "eval", "adjudication_workbook.xlsx")

recs = {json.loads(l)["idx"]: json.loads(l) for l in open(JL, encoding="utf-8")}
verd = json.load(open(VJ, encoding="utf-8"))
# 补 1-160(重复批)为 drop/agree=C
for i in recs:
    k = str(i)
    if k not in verd:
        verd[k] = {"v": "drop", "agree": "C"}
json.dump(verd, open(VJ, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "复核272"
cols = ["序号", "模式", "政策", "题", "强度", "方向", "路径链", "Claude删·理由",
        "DeepSeek留·理由", "原文片段", "Opus预判", "Opus附注", "你的判定", "备注"]
ws.append(cols)
hd = Font(bold=True)
for c in ws[1]:
    c.font = hd
    c.alignment = Alignment(wrap_text=True, vertical="top")

# 排序:flag/keep(分歧) 置顶,再按 模式、序号
def sortkey(i):
    v = verd[str(i)]
    prio = 0 if (v["v"] != "drop" or v.get("agree") != "C") else 1
    return (prio, recs[i]["pattern"], i)

yellow = PatternFill("solid", fgColor="FFF2CC")
orange = PatternFill("solid", fgColor="FCE4D6")
for i in sorted(recs, key=sortkey):
    r = recs[i]
    v = verd[str(i)]
    row = [i, r["pattern"], r["pol"], r["q"], r["str"], r["dir"], r["chain"],
           r["claude_reason"], r["ds_reason"], r["excerpt"],
           v["v"], v.get("note", ""), "", ""]
    ws.append(row)
    flagged = v["v"] != "drop" or v.get("agree") != "C"
    if flagged:
        fill = orange if v["v"] == "flag" else yellow
        for c in ws[ws.max_row]:
            c.fill = fill
for c in ws[1] + tuple(ws.iter_cols(min_row=2)):
    pass
widths = [6, 12, 22, 5, 6, 6, 46, 50, 46, 40, 9, 40, 10, 16]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + j) if j <= 26 else "A" + chr(38 + j)].width = w
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical="top")
ws.freeze_panes = "A2"
wb.save(XLSX)

import collections
flag = sum(1 for k in verd if isinstance(verd[k], dict) and verd[k].get("v") == "flag")
keep = sum(1 for k in verd if isinstance(verd[k], dict) and verd[k].get("v") == "keep")
drop = sum(1 for k in verd if isinstance(verd[k], dict) and verd[k].get("v") == "drop")
print("工作簿 → %s" % XLSX)
print("Opus 预判:drop(同Claude) %d / keep(同DeepSeek) %d / flag(需你定) %d" % (drop, keep, flag))
print("置顶高亮 = 我与Claude不同或存疑的 %d 条;其余 %d 条同Claude=删" % (keep + flag, drop))
